"""One-time: creates Data Hub/Pulse Sheets/Service Budgets.xlsx from
stations.json — one sheet per pulse location, columns Service | Budget,
seeded with the current rates. Also regenerates service_catalog.json, the
name → counting-rule template map build_pulse.py uses to interpret rows."""
import json
import re
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

HERE = Path(__file__).parent
OUT = Path(r"C:\Users\samko\Foxtrot Aviation Services\Data Hub - Documents\Pulse Sheets\Service Budgets.xlsx")

stations = json.loads((HERE / "stations.json").read_text())

norm = lambda s: re.sub(r"[^a-z0-9]", "", s.lower())

# ---- service catalog: normalized name -> spec template ({LOC} placeholder)
catalog = {}
for st_name, cfg in stations.items():
    code = st_name.strip()[:3].upper()
    aircraft_idx = {r - 3 for r in cfg["aircraft_service_rows"]}
    for i, svc in enumerate(cfg["services"]):
        key = norm(svc["name"])
        if key in catalog:
            continue
        entry = {"name": svc["name"], "kind": svc["kind"],
                 "aircraft": (not cfg["fac_only"]) and i in aircraft_idx}
        if svc["kind"] == "count":
            specs = []
            for sp in svc["specs"]:
                crit = [[c, ("{LOC}" if isinstance(v, str)
                             and v.strip().upper() == code else v)]
                        for c, v in sp["criteria"]]
                specs.append({"fn": sp["fn"], "table": sp["table"],
                              "sum_col": sp["sum_col"], "criteria": crit})
            entry["specs"] = specs
        catalog[key] = entry
(HERE / "service_catalog.json").write_text(json.dumps(catalog, indent=2))
print(f"service_catalog.json: {len(catalog)} service templates")
for k, v in catalog.items():
    loc = any("{LOC}" in str(c) for sp in v.get("specs", []) for c in sp["criteria"]) \
        if v["kind"] == "count" else "-"
    print(f"  {v['name']:24} {v['kind']:5} aircraft={v['aircraft']} loc-param={loc}")

# ---- workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)
navy = "FF16314F"
hdr_fill = PatternFill("solid", fgColor=navy)
hdr_font = Font(bold=True, color="FFFFFFFF")
thin = Side(style="thin", color="FFD9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for st_name, cfg in stations.items():
    ws = wb.create_sheet(st_name[:31])
    ws["A1"], ws["B1"] = "Service", "Budget"
    for c in ("A1", "B1"):
        ws[c].fill, ws[c].font, ws[c].border = hdr_fill, hdr_font, border
    r = 2
    for svc in cfg["services"]:
        ws.cell(row=r, column=1, value=svc["name"]).border = border
        cell = ws.cell(row=r, column=2, value=svc.get("rate"))
        cell.border = border
        cell.number_format = "0.00"
        r += 1
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 10
    ws.freeze_panes = "A2"

wb.save(OUT)
print(f"wrote {OUT} ({len(wb.sheetnames)} sheets: {', '.join(wb.sheetnames[:5])}...)")
