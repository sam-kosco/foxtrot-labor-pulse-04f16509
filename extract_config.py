"""One-time extractor: derives each station's pulse-sheet configuration from the
formulas in Live Commercial Pulse Sheet.xlsx and writes stations.json.

Nothing is hand-transcribed: service rows, COUNTIFS/SUMIFS criteria, budget
rates, labor-dist keys, and which rows count as "aircraft" for the weekly
average all come from parsing the week-1 block of each station sheet.
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

WB_PATH = Path(r"C:\Users\samko\Foxtrot Aviation Services\Data Hub - Documents\Pulse Sheets\Live Commercial Pulse Sheet.xlsx")
OUT = Path(__file__).parent / "stations.json"

DATA_SHEETS = {
    "Basic Employee Info", "Envoy Debriefs", "GoJet Debriefs", "Mesa Debriefs",
    "PSA Debriefs", "Breeze Debriefs", "Ultra Debriefs", "This Months Hours",
    "Frontier Debriefs", "Config",
}

# COUNTIFS/SUMIFS criteria pairs: Table[Column],"value"  or  Table[Column],1 / cellref
CRIT_RE = re.compile(r"(\w+)\[([^\]]+)\],\s*(\"[^\"]*\"|[^,()]+)")
FUNC_RE = re.compile(r"(COUNTIFS|SUMIFS)\(([^()]*(?:\([^()]*\)[^()]*)*)\)", re.I)


def parse_specs(formula):
    """Parse =COUNTIFS(...)+COUNTIFS(...) / SUMIFS(...) into evaluable specs.
    Day/Month criteria (cell refs / Config!$B$1) are dropped — the runtime
    supplies them. Returns list of specs, each {fn, table, sum_col, criteria}.
    """
    specs = []
    for fn, body in FUNC_RE.findall(formula):
        spec = {"fn": fn.upper(), "table": None, "sum_col": None, "criteria": []}
        if fn.upper() == "SUMIFS":
            m = re.match(r"\s*(\w+)\[([^\]]+)\]\s*,", body)
            if m:
                spec["table"], spec["sum_col"] = m.group(1), m.group(2)
                body = body[m.end():]
        for tbl, col, val in CRIT_RE.findall(body):
            spec["table"] = spec["table"] or tbl
            val = val.strip()
            if col in ("Day", "Month"):
                continue
            if val.startswith('"'):
                spec["criteria"].append([col, val.strip('"')])
            else:
                try:
                    spec["criteria"].append([col, int(val)])
                except ValueError:
                    spec["criteria"].append([col, val])  # unexpected ref; surface it
        specs.append(spec)
    return specs


def cell_f(ws, coord):
    v = ws[coord].value
    return v if isinstance(v, str) and v.startswith("=") else None


def main():
    wb = openpyxl.load_workbook(WB_PATH, data_only=False)
    wbv = openpyxl.load_workbook(WB_PATH, data_only=True)
    stations = {}

    for name in wb.sheetnames:
        if name in DATA_SHEETS:
            continue
        ws, wsv = wb[name], wbv[name]

        # --- locate week-1 block: first Budgeted/Worked Hours rows after 'Week 1'
        wk1 = bud = wked = None
        for r in range(1, 60):
            a = ws.cell(row=r, column=1).value
            label = a.strip() if isinstance(a, str) else ""
            if label == "Week 1":
                wk1 = r
            elif label == "Budgeted Hours" and wk1 and not bud:
                bud = r
            elif label == "Worked Hours" and wk1 and not wked:
                wked = r
                break
        if not (wk1 and bud and wked):
            print(f"!! {name}: template rows not found (wk1={wk1} bud={bud} wked={wked})",
                  file=sys.stderr)
            continue

        # --- budget table S:T
        budgets = []  # [name, rate] in row order starting S3
        r = 3
        while True:
            s = ws.cell(row=r, column=19).value  # S
            t = ws.cell(row=r, column=20).value  # T
            if s is None:
                break
            budgets.append({"name": str(s).strip(), "rate": t, "row": r})
            r += 1

        # --- service rows between week header and Budgeted Hours
        services = []
        for r in range(wk1 + 1, bud):
            label = ws.cell(row=r, column=1).value
            f = cell_f(ws, f"B{r}")
            if f is None:
                continue
            entry = {"name": str(label).strip(), "sheet_row": r}
            if re.search(r"IF\s*\(\s*OR", f) and "COUNTIFS" not in f and "SUMIFS" not in f:
                # flat daily budget released for elapsed days: IF(OR(...),$T$n,0)
                m = re.search(r"\$T\$(\d+)", f)
                entry["kind"] = "fixed"
                entry["rate_row"] = int(m.group(1)) if m else None
            else:
                entry["kind"] = "count"
                entry["specs"] = parse_specs(f)
            services.append(entry)

        # --- budgeted-hours weights: intended = match service label to budget
        # table label (fixes the IAH row-shift and BDL hardcode bugs). Fall back
        # to positional order among count rows.
        count_rows = [s for s in services if s["kind"] == "count"]
        norm = lambda s: re.sub(r"[^a-z0-9]", "", s.lower())
        bnames = {norm(b["name"]): b for b in budgets}
        for i, s in enumerate(services):
            if s["kind"] == "fixed":
                b = next((x for x in budgets if x["row"] == s.get("rate_row")), None)
                s["rate"] = b["rate"] if b else None
                if b is None and i < len(budgets):
                    s["rate"] = budgets[i]["rate"]
                continue
            b = bnames.get(norm(s["name"]))
            if b is None and i < len(budgets):
                b = budgets[i]
            s["rate"] = b["rate"] if b else 0
        # FAC-only sheets: Budgeted Hours = =B3 (copy of the single fixed row)
        bud_f = cell_f(ws, f"B{bud}") or ""
        fac_only = bool(re.fullmatch(r"=B\d+", bud_f.strip()))

        # --- worked-hours formula: labor dist keys (IAH combines CABIN + FAC)
        wf = cell_f(ws, f"B{wked}") or ""
        labor_keys = re.findall(r"This_Months_Hours\[Labor Dist\],\s*\"([^\"]+)\"", wf)
        salary_keys = re.findall(r"Labor Dist Description\],\s*\"([^\"]+)\"", wf)
        has_salary = "40/7" in wf.replace(" ", "")
        labor_key = labor_keys[0] if labor_keys else None
        salary_key = salary_keys[0] if salary_keys else labor_key

        # --- Average Aircrafts (M4-ish): which service rows count as aircraft
        aircraft_rows = []
        for r in range(wk1, bud + 2):
            f = cell_f(ws, f"M{r}")
            if f and "AVERAGEIF" in f:
                aircraft_rows = [int(x) for x in re.findall(r"B(\d+):H\d+", f)]
                # first ref of each AVERAGEIF pair is the day row; drop it
                aircraft_rows = [x for x in aircraft_rows if x != wk1]
                break
        has_aircraft_kpi = bool(aircraft_rows)

        stations[name] = {
            "labor_keys": labor_keys or ([labor_key] if labor_key else []),
            "salary_keys": salary_keys or ([salary_key] if salary_key else []),
            "labor_key": labor_key,
            "salary_key": salary_key,
            "has_salary": has_salary,
            "fac_only": fac_only,
            "services": [
                {k: v for k, v in s.items() if k not in ("rate_row",)}
                for s in services
            ],
            "aircraft_service_rows": aircraft_rows,
            "budget_table": [[b["name"], b["rate"]] for b in budgets],
            "has_aircraft_kpi": has_aircraft_kpi,
        }
        print(f"{name}: {len(services)} services, labor={labor_key!r}, "
              f"salary={salary_key!r}, fac_only={fac_only}, aircraft_rows={aircraft_rows}")

    OUT.write_text(json.dumps(stations, indent=2))
    print(f"\nwrote {OUT} ({len(stations)} stations)")


if __name__ == "__main__":
    main()
