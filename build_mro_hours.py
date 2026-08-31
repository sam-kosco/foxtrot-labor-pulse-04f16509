"""Builds mro_hours.json — per-day worked hours for each Private/MRO pulse
location, published beside pulse_data.json. The Foxtrot Platform's
Private/MRO Labor calendar consumes it the way the platform homepage
consumes pulse_data.json.

Locations come from Pulse Sheets/Private and MRO Pulse Locations.xlsx: one
worksheet per location, whose A/B columns are a key/value config block
(Labor Distribution, Hourly Goal, Facility Hours, ...). This build reads only
the labor dist(s) and the optional Attribution knob — the price matrix and
budget config are the platform's concern.

Worked hours use the exact contract the commercial pulse uses
(build_pulse.worked_hours: paid-duration punches, holiday/PTO exclusion,
"Not Defined" dist repair, open-shift estimation, 40/7 per active salaried
head), with one deliberate difference: day attribution defaults to the plain
calendar day the punch falls on (Sam, 2026-08-31: "assign work to the day it
is logged on Paylocity" — MRO crews work day shifts, no 12 h shift-back).
A sheet may opt back into the commercial rule with an `Attribution` row
valued "shift".
"""
import json
from datetime import datetime, timedelta

import openpyxl

import build_pulse as bp

MRO_WORKBOOK = "Private and MRO Pulse Locations.xlsx"
# The platform calendar shows 2 weeks back; publish 10 weeks so history is
# there when someone scrolls a schedule conversation backward.
LOOKBACK_DAYS = 70


def load_mro_workbook():
    """One sheet per location. Columns A/B are read as key/values until the
    first blank key, so new knobs slot in without a code change here."""
    wb = openpyxl.load_workbook(bp.BUDGETS_DIR / MRO_WORKBOOK, data_only=True)
    out = {}
    for ws in wb.worksheets:
        kv = {}
        for r in range(1, ws.max_row + 1):
            k = ws.cell(row=r, column=1).value
            if k is None or str(k).strip() == "":
                break
            kv[str(k).strip()] = ws.cell(row=r, column=2).value
        dists = [d.strip() for d in str(kv.get("Labor Distribution") or "").split(",")
                 if d.strip()]
        if not dists:
            raise SystemExit(f"{MRO_WORKBOOK} sheet {ws.title!r} has no "
                             f"'Labor Distribution' value in its A/B block")
        out[ws.title.strip()] = {"dists": dists, "config": kv}
    wb.close()
    if not out:
        raise SystemExit(f"{MRO_WORKBOOK} has no location sheets")
    return out


def assert_no_dist_overlap(mro_locs):
    """A labor dist may feed exactly one pulse. BUR PRIV and DAL PRIV are
    already consumed by the commercial pulse (BUR-SNA, DFW-DAL) — if an MRO
    sheet claimed a dist the commercial config also counts, company-wide
    totals (including the platform homepage stats) would double-count it.
    Hard stop until ownership is decided per dist."""
    base_stations = json.loads((bp.HERE / "stations.json").read_text())
    catalog = json.loads((bp.HERE / "service_catalog.json").read_text())
    catalog.update(json.loads((bp.HERE / "catalog_extras.json").read_text()))
    overrides = json.loads((bp.HERE / "station_overrides.json").read_text())
    commercial = bp.merge_budget_config(
        base_stations, bp.load_budget_workbook(), catalog, overrides)
    owner = {}
    for st, cfg in commercial.items():
        for k in list(cfg["labor_keys"] or []) + list(cfg["salary_keys"] or []):
            owner.setdefault(k.upper(), st)
    clashes = sorted(
        (loc, d, owner[d.upper()])
        for loc, mc in mro_locs.items() for d in mc["dists"]
        if d.upper() in owner)
    if clashes:
        raise SystemExit(
            "labor dist(s) claimed by both pulses:\n"
            + "\n".join(f"  {d!r}: MRO location {loc!r} vs commercial "
                        f"station {st!r}" for loc, d, st in clashes)
            + "\nDecide which pulse owns each dist before publishing "
              "(see MRO_PULSE_PLAN.md in foxtrot-platform).")


def month_slice(hours, col, year, month):
    sel = hours[[d.year == year and d.month == month for d in hours[col]]].copy()
    sel["day"] = [d.day for d in sel[col]]
    return sel


def main():
    mro = load_mro_workbook()
    print(f"{MRO_WORKBOOK}: {len(mro)} location sheet(s): "
          f"{', '.join(sorted(mro))}")
    assert_no_dist_overlap(mro)

    emp = bp.load_employees()
    hours = bp.load_hours(emp)
    start = max(bp.WINDOW_START, bp.TODAY - timedelta(days=LOOKBACK_DAYS))

    series = {loc: {} for loc in mro}
    y, m = start.year, start.month
    while (y, m) <= (bp.TODAY.year, bp.TODAY.month):
        ndays = bp.month_days(y, m)
        hsel_shift = month_slice(hours, "attr_date", y, m)
        hsel_plain = month_slice(hours, "attr_date_plain", y, m)
        for loc, mc in mro.items():
            attribution = str(mc["config"].get("Attribution")
                              or "plain").strip().lower()
            cfg = {"labor_keys": mc["dists"],
                   "facility": attribution != "shift"}
            hourly, est, est_n, sal, worked = bp.worked_hours(
                cfg, hsel_shift, hsel_plain, emp, y, m, ndays)
            for i in range(ndays):
                d = bp.date(y, m, i + 1)
                if start <= d <= bp.TODAY:
                    series[loc][d.isoformat()] = {
                        "worked": worked[i],
                        "hourly": round(hourly[i], 2),
                        "est": est[i],
                        "est_n": est_n[i],
                        "salary_heads": sal[i],
                    }
        m += 1
        if m == 13:
            y, m = y + 1, 1

    for loc in sorted(mro):
        days = series[loc]
        total = sum(v["worked"] for v in days.values())
        print(f"  {loc} ({', '.join(mro[loc]['dists'])}): {len(days)} day(s), "
              f"{total:,.1f} worked h in window")

    data = {
        "generated": datetime.now().strftime("%b %d, %Y %I:%M %p"),
        "window": {"start": start.isoformat(), "end": bp.TODAY.isoformat()},
        "locations": {
            loc: {"dists": mc["dists"], "days": series[loc]}
            for loc, mc in mro.items()
        },
    }
    out = bp.HERE / "mro_hours.json"
    out.write_text(json.dumps(data))
    print(f"wrote mro_hours.json ({out.stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
