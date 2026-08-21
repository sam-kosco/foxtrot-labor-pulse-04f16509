"""Builds pulse_data.json from the same Data Hub sources the Live Commercial
Pulse Sheet queries — the workbook itself is never read.

Replicates: Power Query transforms (per-airline debrief normalization), the
sheet helper columns (punch-in minus 12h day attribution, labor-dist repair,
pay-type lookup), and the sheet formulas (COUNTIFS service counts, budgeted
= rate x count + fixed elapsed-day budgets, worked = hourly punches + 40/7
per active salaried head, weekly averages over elapsed days, variance text).
"""
import json
import os
import re
from calendar import monthrange
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import pandas as pd

HERE = Path(__file__).parent
if os.environ.get("PULSE_DATA_DIR"):
    # CI mode: fetch_sources.py has downloaded everything into one flat dir
    DEBRIEFS = PAYLOCITY = LOCMGMT_DIR = BUDGETS_DIR = Path(os.environ["PULSE_DATA_DIR"])
else:
    DH = Path(r"C:\Users\samko\Foxtrot Aviation Services\Data Hub - Documents")
    DEBRIEFS = DH / "Power Flows" / "Debriefs"
    PAYLOCITY = DH / "Paylocity Reports"
    LOCMGMT_DIR = DH / "Power BI Data Sources"
    BUDGETS_DIR = DH / "Pulse Sheets"
WINDOW_START = date(2026, 5, 1)  # same cutoff the workbook queries hardcode
EXCLUDED_TITLES = {"Director", "Regional Director", "Regional Manager II"}
TODAY = date.today()


def read_table(path, table_name):
    """Read a named Excel table into a DataFrame (matches what Power Query's
    Source{[Item=...,Kind="Table"]} sees)."""
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    for ws in wb.worksheets:
        if table_name in ws.tables:
            ref = ws.tables[table_name].ref
            rows = list(ws[ref])
            header = [c.value for c in rows[0]]
            data = [[c.value for c in r] for r in rows[1:]]
            wb.close()
            return pd.DataFrame(data, columns=header)
    wb.close()
    raise KeyError(f"table {table_name!r} not in {path.name}")


def yes(v):
    return 1 if isinstance(v, str) and v.strip().startswith("Yes") else 0


def not_no(v):
    return 0 if (isinstance(v, str) and v.strip() == "No") else (1 if v else 0)


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return pd.to_datetime(v).date() if v else None


# ---------------------------------------------------------------- debriefs

def load_envoy():
    main = read_table(DEBRIEFS / "Envoy Debriefs.xlsx", "Table1")
    main = pd.DataFrame({
        "date": main["Date"].map(to_date),
        "Location": main["Location"].astype(str).str.strip(),
        "Tail": main["Tail Number"],
        "IHC": main["IHC"].map(yes),
        "RRON": 0,
        "ED1": main["Exterior Detail #1 (ED1)"].map(yes),
        "ED2": main["Exterior Detail #2 (ED2)"].map(yes),
    })
    dfw = read_table(DEBRIEFS / "Envoy Debriefs.xlsx", "DFW_Debriefs")
    dfw = pd.DataFrame({
        "date": dfw["Date"].map(to_date),
        "Location": "DFW",
        "Tail": dfw["Tail"],
        "IHC": dfw["IHC"].map(yes),
        "RRON": dfw["RRON"].map(yes),
        "ED1": dfw["ED1"].map(yes),
        "ED2": dfw["ED2"].map(yes),
    })
    return pd.concat([main, dfw], ignore_index=True)


def load_gojet():
    t = read_table(DEBRIEFS / "GoJet Debriefs.xlsx", "Input")
    return pd.DataFrame({
        "date": t["Date"].map(to_date),
        "Location": t["Location"].astype(str).str.strip(),
        "IHC": t["Interior Heavy Clean (IHC)"].map(not_no),
        "RON": t["RON Clean (RON)"].map(not_no),
        "CE": t["Carpet Extraction (CE)"].map(not_no),
        "ED 1&2": [
            0 if a == "No" and b == "No" else 1
            for a, b in zip(t["Exterior Detail #1 (ED1)"], t["Exterior Detail #2 (ED2)"])
        ],
        "ED 3&4": 0,
    })


def load_mesa():
    t = read_table(DEBRIEFS / "Mesa Debriefs.xlsx", "Input")
    out = pd.DataFrame({"date": t["Date"].map(to_date)})
    for flag, col in [
        ("IHC", "Interior Heavy Clean (IHC)"), ("RON", "RON Clean (RON)"),
        ("EC", "Exterior Clean (EC)"), ("ED", "Exterior Detail (ED)"),
        ("DSC", "Deep Seat Clean (DSC)"), ("CE", "Carpet Extraction (CE)"),
        ("ESS", "Disinfection (ESS)"), ("FDC", "Detailed Flight Deck Clean (Flight Deck)"),
    ]:
        out[flag] = t[col].map(yes) if col in t.columns else 0
    return out


def load_psa():
    t = read_table(DEBRIEFS / "PSA Debriefs.xlsx", "Input")
    svc7 = ["Interior Clean (I)", "Exterior Clean (E)", "Cockpit Cleaning (CC)",
            "Deep Seat Clean (DSC)", "Carpet Extraction (CE)",
            "Exterior Detail (ED1)", "Exterior Detail (ED2)"]
    combined = t[svc7].fillna("").astype(str).agg("".join, axis=1)
    return pd.DataFrame({
        "date": t["Date"].map(to_date),
        "Location": t["Location"].astype(str).str.strip().str[:3],
        "RON": (combined != "No" * 7).astype(int),
        "IC": t["Interior Clean (I)"].map(yes),
        "EC": t["Exterior Clean (E)"].map(yes),
        "CC": t["Cockpit Cleaning (CC)"].map(yes),
        "DSC": t["Deep Seat Clean (DSC)"].map(yes),
        "CE": t["Carpet Extraction (CE)"].map(yes),
        "ED1": t["Exterior Detail (ED1)"].map(yes),
        "ED2": t["Exterior Detail (ED2)"].map(yes),
        "ED3": t["Exterior Detail (ED3)"].isin(["Yes. 900", "Yes. 700"]).astype(int),
        "ED4": t["Exterior Detail (ED4)"].isin(["Yes. 700", "Yes. 900"]).astype(int),
        "Lav": t["Lav Tank Pressure Washing"].map(yes),
    })


def load_breeze():
    t = read_table(DEBRIEFS / "Breeze Debriefs.xlsx", "Table1")
    return pd.DataFrame({
        "date": t["Date"].map(to_date),
        "Location": t["Location"].astype(str).str.strip(),
        "Breeze RON": pd.to_numeric(t["Breeze RON"], errors="coerce").fillna(0).astype(int),
        "Breeze Ultra": pd.to_numeric(t["Breeze Ultra"], errors="coerce").fillna(0).astype(int),
    })


def load_ultra():
    t = read_table(DEBRIEFS / "Ultra Debriefs.xlsx", "Table1")
    out = pd.DataFrame({
        "date": t["Date"].map(to_date),
        "Location": t["Location"].astype(str).str.strip(),
        # One table holds two job types — "Ultra Cleaning" and "Shroud
        # Cleaning". Station specs filter on this so shroud jobs aren't
        # counted as Ultras (the workbook's COUNTIFS filtered Location only).
        "Service": t["Service"].astype(str).str.strip(),
    })
    shroud = next((c for c in t.columns if c and "Shroud" in str(c)), None)
    out["Shroud Count"] = (
        pd.to_numeric(t[shroud], errors="coerce").fillna(0) if shroud else 0
    )
    return out


AA_JOB_TYPES = {"DTC": "AA Turn", "RSTC": "AA Turn", "RON": "AA RON",
                "RRON": "AA RON", "Security": "AA Security",
                "Ultra": "Ultra", "Shroud": "Shroud Cleaning"}


def load_aa():
    t = pd.read_csv(DEBRIEFS / "AA Debriefs.csv", dtype=str)
    t = t[t["Status"].str.strip() == "Completed"]
    svc = t["Job Type"].str.strip().map(AA_JOB_TYPES)
    out = pd.DataFrame({
        "date": pd.to_datetime(t["Job Date"], errors="coerce").dt.date,
        "Station": t["Station"].astype(str).str.strip(),
        "Service": svc,
    })
    return out[out["Service"].notna()]


# APU Wash Sheet2 replaced its Status column with one numeric column per
# service (Aug 2026): 1 = completed, 0.5 = cancelled but half billable,
# 0 = no job. Only a 1 counts toward the pulse.
APU_SERVICES = ["APU Wash", "Landing Gear Bay Wash", "Partial Belly Wash"]


def load_apu():
    t = pd.read_excel(DEBRIEFS / "APU Wash.xlsx", sheet_name="Sheet2")
    t.columns = [str(c).strip() for c in t.columns]
    out = pd.DataFrame({
        "date": pd.to_datetime(t["Date"], errors="coerce").dt.date,
        "Location": t["Location"].astype(str).str.strip(),
    })
    for col in APU_SERVICES:
        if col not in t.columns:
            raise SystemExit(f"APU Wash.xlsx Sheet2 is missing column {col!r} "
                             f"(found {list(t.columns)})")
        out[col] = (pd.to_numeric(t[col], errors="coerce") == 1).astype(int)
    return out


JSX_SERVICES = ["RON", "Interior Detail", "Exterior Detail", "Carpet Extraction",
                "Biohazard"]
# Foxtrot took over the JSX contract on 2026-08-01. Rows dated before that are
# backfilled service history from the previous vendor (they all carry no
# revenue) — they are not work we performed, so they never count.
JSX_START = date(2026, 8, 1)


def load_jsx():
    """JSX Debriefs 'Debriefs' table: one row per tail serviced, with a 0/1
    column per service and the airport in 'Service Location'."""
    t = read_table(DEBRIEFS / "JSX Debriefs.xlsx", "Debriefs")
    out = pd.DataFrame({
        "date": t["Date"].map(to_date),
        "Location": t["Service Location"].astype(str).str.strip().str.upper(),
    })
    for col in JSX_SERVICES:
        if col not in t.columns:
            raise SystemExit(f"JSX Debriefs is missing column {col!r} "
                             f"(found {list(t.columns)})")
        out[col] = (pd.to_numeric(t[col], errors="coerce").fillna(0) == 1).astype(int)
    before = len(out)
    out = out[[d is not None and d >= JSX_START for d in out["date"]]]
    print(f"  JSX: dropped {before - len(out)} pre-{JSX_START} row(s) "
          f"(pre-contract service history)")
    return out


def load_frontier():
    t = read_table(DEBRIEFS / "Frontier Debriefs.xlsx", "Table1")
    t = t[t["Status"] == "Complete"]
    return pd.DataFrame({
        "date": t["Date"].map(to_date),
        "Location": t["Location"].astype(str).str.strip(),
        "Aircraft Type and Service": t["Aircraft Type and Service"],
    })


# ---------------------------------------------------------------- paylocity

NORM = lambda s: re.sub(r"[^a-z0-9]", "", str(s).lower())


def load_budget_workbook():
    """Service Budgets.xlsx: one sheet per location, Service | Budget rows.
    This workbook is the authority on which locations exist, which services
    each has, and the hours-per-job rates."""
    wb = openpyxl.load_workbook(BUDGETS_DIR / "Service Budgets.xlsx", data_only=True)
    out = {}
    for ws in wb.worksheets:
        rows = []
        for r in range(2, ws.max_row + 1):
            svc, rate = ws.cell(row=r, column=1).value, ws.cell(row=r, column=2).value
            if svc is None or str(svc).strip() == "":
                continue
            rows.append((str(svc).strip(), float(rate) if rate is not None else 0.0))
        if rows:
            out[ws.title.strip()] = rows
    wb.close()
    return out


def merge_budget_config(stations, budgets, catalog, overrides):
    """Build the effective per-station config: the budget workbook decides the
    station list, service list, and rates; specs come from station_overrides
    first, then stations.json for pre-existing station+service pairs, then the
    service catalog templates ({LOC} = first 3 letters of the sheet name).
    Unknown service names fall back to a flat daily budget, with a warning."""
    merged = {}
    for st_name, rows in budgets.items():
        code = st_name.strip()[:3].upper()
        base = stations.get(st_name)
        ovr = overrides.get(st_name, {})
        ovr_services = ovr.get("services", {})
        base_by_name = ({NORM(s["name"]): (i, s) for i, s in enumerate(base["services"])}
                        if base else {})
        base_aircraft = ({r - 3 for r in base["aircraft_service_rows"]} if base else set())
        services = []
        for svc_name, rate in rows:
            key = NORM(svc_name)
            if key in ovr_services:
                svc = dict(ovr_services[key])
            elif key in base_by_name:
                i, src = base_by_name[key]
                svc = {k: v for k, v in src.items() if k != "sheet_row"}
                svc["aircraft"] = (not base["fac_only"]) and i in base_aircraft
            elif key in catalog:
                tpl = catalog[key]
                svc = {"kind": tpl["kind"], "aircraft": tpl["aircraft"]}
                if tpl["kind"] == "count":
                    svc["specs"] = [
                        {"fn": sp["fn"], "table": sp["table"], "sum_col": sp["sum_col"],
                         "criteria": [[c, (code if v == "{LOC}" else v)]
                                      for c, v in sp["criteria"]]}
                        for sp in tpl["specs"]]
            else:
                print(f"  !! {st_name}: unknown service {svc_name!r} — "
                      f"treating as flat daily budget")
                svc = {"kind": "fixed", "aircraft": False}
            svc["name"], svc["rate"] = svc_name, rate
            services.append(svc)
        fac_only = all(s["kind"] == "fixed" for s in services)
        if "labor_keys" in ovr:
            labor_keys = ovr["labor_keys"]
            salary_keys = ovr.get("salary_keys", labor_keys)
        elif base:
            labor_keys, salary_keys = base["labor_keys"], base["salary_keys"]
        else:
            k = f"{code} FAC" if "FAC" in st_name.upper() else f"{code} CABIN"
            labor_keys = salary_keys = [k]
            print(f"  new location {st_name!r}: labor dist {k!r} (convention)")
        # Facility stations take plain day attribution (no overnight shift-back).
        # Name test first so a facility that gains a counted service keeps the
        # right treatment; fac_only covers any future non-"FAC" naming.
        facility = st_name.strip().upper().endswith("FAC") or fac_only
        merged[st_name] = {"labor_keys": labor_keys, "salary_keys": salary_keys,
                           "fac_only": fac_only, "facility": facility,
                           "hours_from_first_debrief":
                               bool(ovr.get("hours_from_first_debrief")),
                           "services": services}
    for st in stations:
        if st not in budgets:
            print(f"  !! {st} is in stations.json but has no Service Budgets "
                  f"sheet — dropped from the pulse")
    return merged


def load_managers(station_names):
    """Airport code (first 3 letters, both sides) → managers, per Location
    Management.csv. A station lands in every matching manager's group."""
    m = pd.read_csv(LOCMGMT_DIR / "Location Management.csv", dtype=str,
                    encoding="utf-8-sig")
    by_code = {}
    for loc, mgr in zip(m["Location"], m["Manager"]):
        if pd.isna(loc) or pd.isna(mgr):
            continue
        by_code.setdefault(loc.strip()[:3].upper(), set()).add(mgr.strip())
    return {st: sorted(by_code.get(st.strip()[:3].upper(), set()))
            for st in station_names}


def load_employees():
    e = pd.read_csv(PAYLOCITY / "Basic Employee Info.csv", dtype=str, encoding="cp1252")
    e = e[~e["Job Title"].isin(EXCLUDED_TITLES)].copy()
    term = pd.to_datetime(e["Termination Date"], errors="coerce")
    active = e["Employee Status Description"].str.strip() == "Active"
    e["last_day"] = [
        TODAY if a else (t.date() if pd.notna(t) else None)
        for a, t in zip(active, term)
    ]
    e["labor_dist"] = e["Labor Dist Description"].fillna("").str.strip()
    e["pay_type"] = e["Pay Type Code"].fillna("").str.strip()
    e["id"] = e["Employee Id"].str.strip()
    return e


def hours_file_updated():
    """When This Years Hours.csv was last regenerated (naive local/Eastern).
    CI: SharePoint lastModifiedDateTime captured by fetch_sources.py;
    local: file mtime."""
    meta = PAYLOCITY / "sources_meta.json"
    if meta.exists():
        import json as _json
        ts = _json.loads(meta.read_text()).get("This Years Hours.csv")
        if ts:
            from zoneinfo import ZoneInfo
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            return dt.astimezone(ZoneInfo("America/New_York")).replace(tzinfo=None)
    return datetime.fromtimestamp((PAYLOCITY / "This Years Hours.csv").stat().st_mtime)


def load_hours(emp):
    h = pd.read_csv(PAYLOCITY / "This Years Hours.csv", dtype=str, encoding="cp1252")
    # Paylocity renamed this column in the Aug 2026 export; accept either name
    if "Labor Dist Name" not in h.columns and "Cost Center 2 Name" in h.columns:
        h = h.rename(columns={"Cost Center 2 Name": "Labor Dist Name"})
    h["work_date"] = pd.to_datetime(h["Work Date"], errors="coerce").dt.date
    h = h[(h["work_date"] >= WINDOW_START) & (h["work_date"] <= TODAY)].copy()
    h["hours"] = pd.to_numeric(h["Summation of Paid Duration (hours)"], errors="coerce").fillna(0)
    punch = pd.to_datetime(h["Punch In Time"], errors="coerce")
    punched_out = h["Punch Out Time"].fillna("").str.strip() != ""

    # Holiday/PTO credits: punch-in placeholder, no punch-out, but paid hours
    # already posted. Excluded from worked hours (owner decision, Aug 2026).
    holiday = punch.notna() & ~punched_out & (h["hours"] > 0)
    h = h[~holiday].copy()
    punch = punch[~holiday]
    punched_out = punched_out[~holiday]
    print(f"  holiday/PTO rows excluded: {int(holiday.sum())}")

    # Two day attributions, chosen per station in build_month():
    #  attr_date       — workbook helper DAY/MONTH(punch-in − 12h). Commercial
    #                    stations work overnight, so a shift credits the day it
    #                    started.
    #  attr_date_plain — the calendar day the punch actually falls in. Facility
    #                    stations don't run overnight shifts, so shifting back
    #                    12 h would push a normal day shift onto the day before
    #                    (owner decision, Aug 2026).
    # Rows with no punch-in fall back to Work Date in both.
    shifted = punch - timedelta(hours=12)
    h["attr_date"] = [
        s.date() if pd.notna(s) else w for s, w in zip(shifted, h["work_date"])
    ]
    h["attr_date_plain"] = [
        p.date() if pd.notna(p) else w for p, w in zip(punch, h["work_date"])
    ]

    # Shifts still in progress (punch-in, no punch-out, 0 paid hours) at the
    # moment the CSV was generated: estimate the day's hours from history —
    # employee's median closed-shift duration, else location median, else 8 h,
    # capped at 16. Only shifts that STARTED in the 18 h before the file was
    # generated qualify (excludes future/scheduled placeholder punches and
    # long-forgotten open punches). Corrected by real hours at next refresh.
    updated = hours_file_updated()
    age_h = (updated - punch).dt.total_seconds() / 3600
    open_shift = punch.notna() & ~punched_out & (h["hours"] == 0) \
        & (age_h >= 0) & (age_h <= 18)
    closed = h[punched_out & (h["hours"] > 0)]
    emp_med = closed.groupby(closed["Employee Id"].str.strip())["hours"].median()
    dist_med = closed.groupby(closed["Labor Dist Name"].fillna("").str.strip().str.upper())["hours"].median()
    def estimate(row):
        e = emp_med.get(str(row["Employee Id"]).strip())
        if pd.isna(e) or e is None:
            e = dist_med.get(str(row["Labor Dist Name"] or "").strip().upper())
        if pd.isna(e) or e is None:
            e = 8.0
        return min(float(e), 16.0)
    h["est"] = 0.0
    h.loc[open_shift, "est"] = h[open_shift].apply(estimate, axis=1)
    print(f"  open shifts estimated: {int(open_shift.sum())} "
          f"(~{h['est'].sum():.0f} h; file updated {updated:%b %d %I:%M %p})")
    id2dist = dict(zip(emp["id"], emp["labor_dist"]))
    id2pay = dict(zip(emp["id"], emp["pay_type"]))
    eid = h["Employee Id"].str.strip()
    raw = h["Labor Dist Name"].fillna("").str.strip()
    h["labor_dist"] = [
        id2dist.get(i, "") if r == "Not Defined" else r for i, r in zip(eid, raw)
    ]
    h["pay_type"] = eid.map(id2pay).fillna("")
    return h


# ---------------------------------------------------------------- compute

def spec_mask(df, spec):
    m = pd.Series(True, index=df.index)
    for col, val in spec["criteria"]:
        if col not in df.columns:
            return pd.Series(False, index=df.index)
        if isinstance(val, str):
            m &= df[col].astype(str).str.strip().str.upper() == val.strip().upper()
        else:
            m &= df[col] == val
    return m


def month_days(year, month):
    return monthrange(year, month)[1]


DOW_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def dow_daily_rates(svc):
    """Per-weekday budget for a fixed service carrying a `dow_shape` override.

    Some facilities don't staff evenly across the week (FLL FAC works ~2 h on
    Saturday and nothing on Sunday), so a flat daily rate misstates every day
    even when the week is right. `dow_shape` says how a week is distributed.

    The Service Budgets workbook stays authoritative: its rate is per-day, so
    a week is worth rate x 7, and the shape is rescaled to that total. Editing
    the workbook rate therefore still moves the budget instead of being
    silently ignored. Returns a list indexed by date.weekday(), or None to
    keep the flat spread.
    """
    shape = svc.get("dow_shape")
    if not shape:
        return None
    unknown = set(shape) - set(DOW_NAMES)
    if unknown:
        raise SystemExit(f"dow_shape for {svc.get('name')!r} has unknown "
                         f"day(s) {sorted(unknown)}; use {DOW_NAMES}")
    vals = [float(shape.get(n, 0)) for n in DOW_NAMES]
    total = sum(vals)
    if total <= 0:
        return None
    scale = (float(svc.get("rate") or 0) * 7) / total
    return [v * scale for v in vals]


def aa_fill_plan(aa_full):
    """AA's feed doesn't update automatically (runs through AA's internal
    system), so past dates can be entirely absent. For each station, any past
    date with no AA rows gets the 7-day per-service average taken from the 7
    days ending at the most recent listed date before it (owner decision,
    Aug 2026). Returns {station: {date: {service: avg}}}."""
    plans = {}
    for stn, grp in aa_full.groupby("Station"):
        dates = {d for d in grp["date"] if d is not None}
        if not dates:
            continue
        fills = {}
        d = max(min(dates), WINDOW_START)
        while d < TODAY:
            if d not in dates:
                prior = [x for x in dates if x < d]
                if prior:
                    ref = max(prior)
                    win = {ref - timedelta(days=i) for i in range(7)}
                    sel = grp[grp["date"].isin(win)]
                    fills[d] = {svc: round(float(n) / 7, 2)
                                for svc, n in sel.groupby("Service").size().items()}
            d += timedelta(days=1)
        if fills:
            plans[stn] = fills
            print(f"  AA gap-fill {stn}: {len(fills)} missing day(s) "
                  f"filled with 7-day averages")
    return plans


def first_debrief_dates(stations, tables):
    """Earliest date each flagged station has a counted service on.

    Stations carrying `hours_from_first_debrief` only start counting labor from
    their first real debrief: the hours before that are implementation and
    training for a contract that hadn't started, and shouldn't score against
    the station. Derived from the debrief data (not a hardcoded date) so it
    corrects itself as debriefs land or are backfilled.
    """
    out = {}
    for st_name, cfg in stations.items():
        if not cfg.get("hours_from_first_debrief"):
            continue
        best = None
        for svc in cfg["services"]:
            for spec in svc.get("specs", []):
                t = tables.get(spec["table"])
                if t is None or t.empty:
                    continue
                m = spec_mask(t, spec)
                if spec["fn"] == "SUMIFS":
                    m &= pd.to_numeric(t[spec["sum_col"]],
                                       errors="coerce").fillna(0) > 0
                dates = [d for d in t.loc[m, "date"] if d is not None]
                if dates and (best is None or min(dates) < best):
                    best = min(dates)
        out[st_name] = best
        if best:
            print(f"  {st_name}: labor counted from first debrief {best}")
        else:
            print(f"  !! {st_name}: no debriefs at all — all labor suppressed")
    return out


def build_month(year, month, stations, tables, hours, emp, aa_plans, hours_start):
    ndays = month_days(year, month)
    is_current = (year, month) == (TODAY.year, TODAY.month)
    is_past = date(year, month, 1) < date(TODAY.year, TODAY.month, 1)
    # elapsed-day cutoff, exactly as the sheets do it: past month => all days,
    # current month => days strictly before today
    cutoff = ndays + 1 if is_past else (TODAY.day if is_current else 0)

    # pre-slice each debrief table to this month
    month_tbl = {}
    for tname, df in tables.items():
        sel = df[[d is not None and d.year == year and d.month == month
                  for d in df["date"]]].copy()
        sel["day"] = [d.day for d in sel["date"]]
        month_tbl[tname] = sel

    def month_slice(col):
        sel = hours[[d.year == year and d.month == month for d in hours[col]]].copy()
        sel["day"] = [d.day for d in sel[col]]
        return sel

    hsel_shift = month_slice("attr_date")        # commercial stations
    hsel_plain = month_slice("attr_date_plain")  # facility stations

    out = {}
    for st_name, cfg in stations.items():
        days = list(range(1, ndays + 1))
        code = st_name.strip()[:3].upper()
        st_fills = aa_plans.get(code, {})
        aa_est_days = set()
        svc_rows = []
        budgeted = [0.0] * ndays
        for svc in cfg["services"]:
            vals = []
            # AA-fed service? note its Service criterion for gap-filling
            aa_svc = next((v for sp in svc.get("specs", [])
                           if sp["table"] == "AA_Debriefs"
                           for c, v in sp["criteria"] if c == "Service"), None)
            svc_est_days = []
            if svc["kind"] == "fixed":
                daily = dow_daily_rates(svc)
                for d in days:
                    elapsed_day = is_past or d < cutoff
                    if not elapsed_day:
                        vals.append(0)
                        continue
                    v = (daily[date(year, month, d).weekday()] if daily
                         else svc["rate"])
                    vals.append(round(v or 0, 2))
            else:
                for d in days:
                    total = 0
                    for spec in svc["specs"]:
                        t = month_tbl.get(spec["table"])
                        if t is None or t.empty:
                            continue
                        m = spec_mask(t, spec) & (t["day"] == d)
                        if spec["fn"] == "SUMIFS":
                            total += float(t.loc[m, spec["sum_col"]].sum())
                        else:
                            total += int(m.sum())
                    if aa_svc is not None:
                        fill = st_fills.get(date(year, month, d))
                        if fill is not None:
                            total = fill.get(aa_svc, 0)
                            svc_est_days.append(d)
                            aa_est_days.add(d)
                    vals.append(total)
            rate = svc.get("rate") or 0
            for i, v in enumerate(vals):
                budgeted[i] += (v if svc["kind"] == "fixed" else v * rate)
            svc_rows.append({"name": svc["name"], "kind": svc["kind"],
                             "rate": rate, "days": vals,
                             "aircraft": svc.get("aircraft", False),
                             "est_days": svc_est_days})

        # worked hours: hourly punches + salaried imputation, per day
        keys = [k.upper() for k in (cfg.get("labor_keys") or [cfg["labor_key"]]) if k]
        skeys = [k.upper() for k in (cfg.get("salary_keys") or keys) if k] or keys
        hourly = [0.0] * ndays
        est = [0.0] * ndays
        est_n = [0] * ndays
        hsel = hsel_plain if cfg.get("facility") else hsel_shift
        st_h = hsel[(hsel["labor_dist"].str.upper().isin(keys))
                    & (hsel["pay_type"] == "Hourly")]
        for d, v in st_h.groupby("day")["hours"].sum().items():
            if 1 <= d <= ndays:
                hourly[d - 1] = float(v)
        openrows = st_h[st_h["est"] > 0]
        for d, v in openrows.groupby("day")["est"].sum().items():
            if 1 <= d <= ndays:
                est[d - 1] = round(float(v), 2)
        for d, n in openrows.groupby("day").size().items():
            if 1 <= d <= ndays:
                est_n[d - 1] = int(n)
        sal_emp = emp[(emp["pay_type"] == "Salary")
                      & (emp["labor_dist"].str.upper().isin(skeys))]
        sal_counts = []
        for d in days:
            cut = date(year, month, d)
            n = sum(1 for ld in sal_emp["last_day"] if ld and ld > cut)
            sal_counts.append(n)
        # Pre-launch labor (before this station's first debrief) is
        # implementation/training — zero it out rather than score it.
        pre_launch = []
        if st_name in hours_start:
            start = hours_start[st_name]
            for i, dnum in enumerate(days):
                if start is None or date(year, month, dnum) < start:
                    if hourly[i] or est[i] or sal_counts[i]:
                        pre_launch.append(dnum)
                    hourly[i] = 0.0
                    est[i] = 0.0
                    est_n[i] = 0
                    sal_counts[i] = 0

        worked = [round(h + e + n * 40 / 7, 2)
                  for h, e, n in zip(hourly, est, sal_counts)]

        # weekly blocks + stats (avg over elapsed days only)
        weeks = []
        for w in range(4):
            lo = w * 7 + 1
            hi = min(lo + 6, ndays) if w < 3 else ndays
            drange = list(range(lo, hi + 1))
            elapsed = [d for d in drange if is_past or d < cutoff]
            def avg(series):
                pts = [series[d - 1] for d in elapsed]
                return round(sum(pts) / len(pts), 2) if pts else None
            ac = None
            if not cfg["fac_only"]:
                acc = 0.0
                got = False
                for sr in svc_rows:
                    if not sr["aircraft"]:
                        continue
                    a = avg(sr["days"])
                    if a is not None:
                        acc += a
                        got = True
                ac = round(acc, 2) if got else None
            ab, aw = avg(budgeted), avg(worked)
            if ab and aw is not None and ab != 0:
                ratio = aw / ab
                pct = round(abs(1 - ratio), 2) * 100
                var = {"pct": round(pct), "dir": "Over Budget" if ratio > 1 else "Under Budget"}
            else:
                var = None
            weeks.append({"days": drange, "avg_aircraft": ac,
                          "avg_budgeted": ab, "avg_worked": aw, "variance": var})

        mtd_days = [d for d in days if is_past or d < cutoff]
        mtd_b = round(sum(budgeted[d - 1] for d in mtd_days), 1)
        mtd_w = round(sum(worked[d - 1] for d in mtd_days), 1)
        out[st_name] = {
            "services": svc_rows,
            "budgeted": [round(b, 2) for b in budgeted],
            "worked": worked,
            "hourly": [round(h, 2) for h in hourly],
            "est": est,
            "est_n": est_n,
            "aa_est_days": sorted(aa_est_days),
            "salary_heads": sal_counts,
            "weeks": weeks,
            "mtd": {"budgeted": mtd_b, "worked": mtd_w,
                    "variance": (round(abs(1 - mtd_w / mtd_b) * 100)
                                 if mtd_b else None),
                    "over": mtd_w > mtd_b if mtd_b else None},
            "elapsed_through": (cutoff - 1) if not is_past else ndays,
        }
    return {"year": year, "month": month, "ndays": ndays,
            "label": date(year, month, 1).strftime("%B %Y"), "stations": out}


def check_sources_present():
    """Fail fast and by name if a source the build needs wasn't fetched.
    Guards against the fetch list and the build drifting apart."""
    from sources import SOURCE_NAMES
    roots = {"This Years Hours.csv": PAYLOCITY, "Basic Employee Info.csv": PAYLOCITY,
             "Location Management.csv": LOCMGMT_DIR, "Service Budgets.xlsx": BUDGETS_DIR}
    missing = [n for n in SOURCE_NAMES if not (roots.get(n, DEBRIEFS) / n).exists()]
    if missing:
        raise SystemExit(
            "missing source file(s): " + ", ".join(missing) +
            "\nIf this is CI, check that sources.py lists them and "
            "fetch_sources.py downloaded them.")


def main():
    check_sources_present()
    base_stations = json.loads((HERE / "stations.json").read_text())
    catalog = json.loads((HERE / "service_catalog.json").read_text())
    catalog.update(json.loads((HERE / "catalog_extras.json").read_text()))
    overrides = json.loads((HERE / "station_overrides.json").read_text())
    budgets = load_budget_workbook()
    print(f"Service Budgets.xlsx: {len(budgets)} location sheets")
    stations = merge_budget_config(base_stations, budgets, catalog, overrides)
    print("loading sources...")
    tables = {
        "Envoy_Debriefs": load_envoy(),
        "GoJet_Debriefs": load_gojet(),
        "Mesa_Debriefs": load_mesa(),
        "PSA_Debriefs": load_psa(),
        "Breeze_Debriefs": load_breeze(),
        "Ultra_Debriefs": load_ultra(),
        "Frontier_Debriefs": load_frontier(),
        "AA_Debriefs": load_aa(),
        "APU_Wash": load_apu(),
        "JSX_Debriefs": load_jsx(),
    }
    for k, v in tables.items():
        print(f"  {k}: {len(v)} rows")
    emp = load_employees()
    hours = load_hours(emp)
    print(f"  employees: {len(emp)}, punch rows in window: {len(hours)}")

    aa_plans = aa_fill_plan(tables["AA_Debriefs"])
    hours_start = first_debrief_dates(stations, tables)
    months = []
    y, m = WINDOW_START.year, WINDOW_START.month
    while (y, m) <= (TODAY.year, TODAY.month):
        months.append(build_month(y, m, stations, tables, hours, emp, aa_plans,
                                  hours_start))
        m += 1
        if m == 13:
            y, m = y + 1, 1
    data = {
        "generated": datetime.now().strftime("%b %d, %Y %I:%M %p"),
        "months": months,
        "station_names": list(stations.keys()),
        "managers": load_managers(stations.keys()),
    }
    (HERE / "pulse_data.json").write_text(json.dumps(data))
    print(f"wrote pulse_data.json ({(HERE / 'pulse_data.json').stat().st_size // 1024} KB, "
          f"{len(months)} months)")

    template = (HERE / "template.html").read_text(encoding="utf-8")
    html = template.replace("/*__DATA__*/", json.dumps(data))
    (HERE / "index.html").write_text(html, encoding="utf-8")
    print(f"wrote index.html ({(HERE / 'index.html').stat().st_size // 1024} KB)")


if __name__ == "__main__":
    main()
